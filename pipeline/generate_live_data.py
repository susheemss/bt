"""
generate_live_data.py
Reads data.xlsx -> writes live_data.js

New format (demand-only, no inventory sheet):
  One sheet, header on row 1:
  SKU Name | Store Name | Week Start Date | Baseline (units) | Promo Units | Total Demand

Stores and SKUs are fully dynamic -- whatever names appear in the file.
There is no Store ID / SKU ID column, so the store/SKU name IS the identifier
(slugified for use as a JS object key). There is no inventory data (on-hand,
in-transit, safety stock, ROP, net requirement) and no Weather/Festival
signal columns -- only Promo.

Run this script whenever the data team drops a new data.xlsx.
Output file is loaded by replenishment-cockpit-v5.html automatically.
"""
import sys, json, re
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from pathlib import Path
from collections import defaultdict
from datetime import datetime, date

EXCEL_PATH  = Path(__file__).parent.parent / 'data.xlsx'
OUTPUT_PATH = Path(__file__).parent.parent / 'live_data.js'

REQUIRED_COLS = ['SKU Name', 'Store Name', 'Week Start Date', 'Baseline (units)', 'Promo Units']


def slugify(name: str) -> str:
    s = re.sub(r'[^a-z0-9]+', '-', name.strip().lower()).strip('-')
    return s or 'store'


def normalize_date(v):
    if isinstance(v, (datetime, date)):
        return v.isoformat()[:10]
    return str(v).strip()


# ── FIND THE DATA SHEET ─────────────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

def read_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h else '' for h in rows[0]]
    data = []
    for row in rows[1:]:
        if any(v is not None for v in row):
            data.append(dict(zip(headers, row)))
    return data

demand_rows = None
for sheet_name in wb.sheetnames:
    candidate = read_sheet(wb[sheet_name])
    if candidate and all(col in candidate[0] for col in REQUIRED_COLS):
        demand_rows = candidate
        break

if demand_rows is None:
    raise SystemExit(f'No sheet in {EXCEL_PATH.name} has the expected columns: {REQUIRED_COLS}')

# ── AGGREGATE: store -> week_index -> {baseline, promo} ────────────────────
week_dates_by_store = defaultdict(set)
for r in demand_rows:
    store = str(r.get('Store Name') or '').strip()
    wk    = r.get('Week Start Date')
    if store and wk is not None:
        week_dates_by_store[store].add(normalize_date(wk))

sorted_weeks = {store: sorted(dates) for store, dates in week_dates_by_store.items()}

store_week = defaultdict(lambda: defaultdict(lambda: {'baseline': 0.0, 'promo': 0.0}))
store_sku_week = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: {'baseline': 0.0, 'promo': 0.0})))
store_sku_names = defaultdict(dict)  # store -> sku_id -> display name

for r in demand_rows:
    store = str(r.get('Store Name') or '').strip()
    sku   = str(r.get('SKU Name') or '').strip()
    wk    = r.get('Week Start Date')
    if not store or not sku or wk is None:
        continue
    wi = sorted_weeks[store].index(normalize_date(wk))
    bl = float(r.get('Baseline (units)') or 0)
    pr = float(r.get('Promo Units') or 0)

    store_week[store][wi]['baseline'] += bl
    store_week[store][wi]['promo']    += pr

    sku_id = slugify(sku)
    store_sku_names[store][sku_id] = sku
    store_sku_week[store][sku_id][wi]['baseline'] += bl
    store_sku_week[store][sku_id][wi]['promo']    += pr

# ── BUILD LIVE_STORES OUTPUT ────────────────────────────────────────────────
LIVE_STORES = {}

for store, weeks in sorted_weeks.items():
    store_id = slugify(store)
    num_weeks = len(weeks)

    forecast_arr, sensed_arr = [], []
    for w in range(num_weeks):
        d = store_week[store][w]
        bl = round(d['baseline'], 2)
        sensed = round(d['baseline'] + d['promo'], 2)
        forecast_arr.append(bl)
        sensed_arr.append(sensed)

    w0 = store_week[store][0]
    bl0 = w0['baseline'] if w0['baseline'] else 1
    promo_pct = round(w0['promo'] / bl0 * 100, 1)
    signals = [
        {'t': 'Promo',    'pct': promo_pct, 'on': w0['promo'] > 0},
        {'t': 'Weather',  'pct': 0, 'on': False},
        {'t': 'Festival', 'pct': 0, 'on': False},
        {'t': 'Trend',    'pct': 0, 'on': False},
        {'t': 'Event',    'pct': 0, 'on': False},
    ]
    uplift = promo_pct

    skus = []
    for sku_id, sku_name in store_sku_names[store].items():
        w0_sku = store_sku_week[store][sku_id].get(0, {'baseline': 0})
        skus.append({'id': sku_id, 'name': sku_name, 'dmd': round(w0_sku['baseline'], 2)})

    LIVE_STORES[store_id] = {
        'name':     store,
        'forecast': forecast_arr,
        'sensed':   sensed_arr,
        'signals':  signals,
        'uplift':   uplift,
        'skus':     skus,
        'weeks':    num_weeks,
    }

# ── WRITE live_data.js ──────────────────────────────────────────────────────
js_content = f"""/* AUTO-GENERATED — do not edit manually.
   Source: data.xlsx (demand-only format: SKU Name, Store Name, Week Start Date,
   Baseline (units), Promo Units, Total Demand). No inventory data yet.
   Run pipeline/generate_live_data.py to refresh.
*/
window.LIVE_STORES = {json.dumps(LIVE_STORES, indent=2)};
"""

OUTPUT_PATH.write_text(js_content, encoding='utf-8')
print(f'Written: {OUTPUT_PATH}')
for k, v in LIVE_STORES.items():
    active_sigs = [s['t'] for s in v['signals'] if s['on']]
    print(f"  Store {k} ({v['name']}): weeks={v['weeks']}, uplift={v['uplift']}%, "
          f"signals={active_sigs}, SKUs={len(v['skus'])}")
